# -*- coding: utf-8 -*-

# Programmed  by Mingyu 'Ozone' CUI | DCOE, Sonepar APAC
# Revision 1: Add function to delete unmatched lines in 365 (.xls) template

import warnings
warnings.filterwarnings("ignore")

import tkinter as tk
import windnd
from datetime import datetime
import os
import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageTk
import xlwings as xw

photo = Image.open('Hagemeyer.png')
print("Warnings surpressed.")

# -----[GUI]-----
def dragged_files(files):
    global order_dic
    global temp_dic
    global tp
    global counter
    global excel_temp
    
    msg = ''.join((item.decode('gbk') for item in files))
    fp = msg.split('\n')[0]
    text0.insert('insert', fp)
    
    if counter == 1:
        # openpyxl numbers sheet columns & rows starting from 1.
        model = 5 # Column 'E'
        amount = 7 # Column 'G'
        delivery = 13 # Column 'M'
        excel = load_workbook(fp)
        sheet = excel.worksheets[0]
        for r in range(1,sheet.max_row+1):
            l = []
            i = r - 2
            if r ==1:
                continue
            else:
                k = sheet.cell(row=r,column=model).value
                k = k.replace(' ', '').replace('-', '').replace('\n', '')
                l.append(sheet.cell(row=r,column=amount).value)
                #deliv = datetime.strptime(sheet.cell(row=r,column=delivery).value, '%m/%d/%Y') # String type value in 'delivery date' column
                deliv = sheet.cell(row=r,column=delivery).value
                l.append(deliv)
                order_dic[k] = l
        print('Total Order lines extracted: ', len(order_dic))
    
    if counter == 0:
        # xlrd numbers sheet columns & rows starting from 0.
        model = 1 # Column 'B'
        tp = fp
        excel_temp = xlrd.open_workbook(fp, formatting_info=True) # Read Excel file with existing formats maintained.
        sheet = excel_temp.sheets()[0]
        nrows = sheet.nrows
        for i in range(1, nrows): # SKIP COLUMN NAME ROW (#1)
            c = sheet.cell_value(i, model)
            c = c.replace(' ', '').replace('-', '').replace('\n', '')
            temp_dic[c] = None
        print('Template loaded from: ', tp)
    
    return

class interface():
    def __init__(self, master, bg, ttl, btn_cap):
        global boo
        
        self.master = master
        self.master.config(bg=bg)
        self.master.title(ttl)
        self.master.geometry('600x400')
        self.interface = tk.Frame(self.master)
        self.interface.pack()
        self.label = tk.Label(self.interface, image = hag, text = 'Developed by:\n   Hagemeyer | DCOE Sonepar\n Yiping LU | Ozone CUI', font=('Arial', 6), compound='left', width=800, height=38).pack(side='bottom', anchor='e')
        
        
        if counter == 0:
            btn_cmd = self.process
        elif counter == 1:
            boo = 1
            btn_cmd = self.next
        else:
            btn_cmd = self.next
        
        self.label = tk.Label(self.interface, text = '请按窗口标题提示拖放Excel文件，等待下方显示文件路径后点击按钮。', fg = '#10387d', bg='white', font=('Arial', 10), width=400, height=3)
        self.label.pack()
        
        tk.Button(self.interface,text=btn_cap, font=('Arial', 12), command=btn_cmd, width=20, height=2).pack(side='bottom')
    
    def next(self,):
        global boo
        global counter
        global total
        self.interface.destroy()
        text0.delete(0.0, tk.END)
        counter -= 1
        num = total - counter
        interface(self.master, colors[0], sheet_names[num], button_caps[boo])
    
    def process(self,):
        global order_dic
        global temp_dic
        global tp
        global excel_temp
        
        # xlwt numbers sheet columns & rows starting from 0.
        model = 1 # Column 'B'
        amount = 4 # Column 'E'
        delivery = 6 # Column 'G'
        line = 0
        template = copy(excel_temp)
        ws = template.get_sheet(0)
        sheet = excel_temp.sheets()[0]
        
        for ko in list(order_dic.keys()):
            if ko in temp_dic:
                temp_dic[ko] = order_dic[ko]
                print(ko, ': ', temp_dic[ko])
                line += 1
        
        style = xlwt.XFStyle()
        style.num_format_str = 'YYYY-MM-DD'
        nrows = sheet.nrows
        for i in range(1, nrows):
            c = sheet.cell_value(i, model)
            if c in order_dic:
                ws.write(i, amount, order_dic[c][0])
                ws.write(i, delivery, order_dic[c][1], style)
        
        template.save(tp)
        print('Total Order lines matched & written to Template: ', line)
        
        # DELETE UNMATCHED TEMPLATE LINES.
        app = xw.App(visible=True, add_book=False)
        app.display_alerts=False   #警告提示，不显示Excel消息框
        app.screen_updating=False  #关闭屏幕更新,可加快宏的执行速度
        
        wb = app.books.open(tp)
        sheet =wb.sheets[0]
        
        rows = sheet.api.UsedRange.Rows.count #总行数
        cols = sheet.api.UsedRange.columns.count  #总列数
        
        c = 'E2:E'+str(rows) #整列的范围
        a_range = sheet.range(str(c))  # 得到range对象,就是amount所在的列
        
        #将range中每行对象存放到列表中并倒序（倒序是因为删除行后，后面的行位置都-1，从后向前删除就不会存在这个问题了）
        cell_list = []
        for cell in a_range:
            cell_list.append(cell)
        cell_list.reverse()
        
        #删除行操作
        delRowNum = 0
        for cell in cell_list:
            if cell.value == '':
                cell_to_del = cell.address
                sheet.range(cell_to_del).api.EntireRow.Delete()
                delRowNum += 1
        
        wb.save(tp)
        rows = sheet.api.UsedRange.Rows.count #总行数
        print('Unused rows deleted: ', delRowNum, '\nTotal final rows in template: ', rows)
        
        wb.close()
        app.quit() 
        
        self.master.destroy()


# INSTANTIATE GUI

global counter
global total
global order_dic # CREATE A DICTIONARY FOR ALL ORDER LINE ITEMS.
global temp_dic # CREATE A DICTIONARY TO CARRY VALUES TO PASS.
global boo
global tp
global excel_temp
sheet_names = ['导入数据：AX订单', '目标表格：365模板']
colors = ['#006699']
counter = len(sheet_names) - 1
total  = len(sheet_names) - 1
print('GUI loaded: Ready for document process.')
order_dic = {}
temp_dic = {}
button_caps = ['下一步 >>', '处理数据']
boo = 0
tp = ''

if __name__ == '__main__':
    root = tk.Tk()
    text0 = tk.Text(root, width=500, height=1, font=('Arial', 12))
    text0.pack(side='bottom')
    hag = ImageTk.PhotoImage(photo.resize((180, 30)))
    interface(root, colors[0], sheet_names[0], button_caps[0])
    windnd.hook_dropfiles(root, func=dragged_files)
    root.mainloop()